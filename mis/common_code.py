import logging
import os
from urllib import parse
from io import BytesIO

import xlrd
from django.db.models import ProtectedError
from django.db.transaction import atomic
from django.http import HttpResponse
from rest_framework import mixins, status, serializers
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.generics import GenericAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from mis import settings
from openpyxl import load_workbook, cell
from openpyxl.utils import get_column_letter

from user.serializers import UserOperationLogSerializer

error_logger = logging.getLogger('error_log')

# 通用导出视图
from user.models import User, Permissions


class CommonExportListMixin(mixins.ListModelMixin):
    SHEET_NAME = 'sheet1'
    TEMPLATE_FILE = 'xlsx_template/example.xlsx'
    VALUES_FIELDS = []

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        export = self.request.query_params.get('export')
        all_flag = self.request.query_params.get('all')
        if export:
            data = self.get_serializer(queryset, many=True).data
            return gen_template_response(self.EXPORT_FIELDS_DICT, data, self.FILE_NAME,
                                         self.SHEET_NAME, self.TEMPLATE_FILE)
        if all_flag and self.VALUES_FIELDS:
            return Response(queryset.values(*self.VALUES_FIELDS))
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


# 通用批量删除、启用/禁用视图
class CommonBatchDestroyView(GenericAPIView):

    @atomic()
    @action(methods=['post'], detail=False, permission_classes=[IsAuthenticated], url_path='batch-destroy',
            url_name='batch-destroy')
    def batch_destroy(self, request):
        obj_ids = self.request.data.get('obj_ids')
        try:
            self.get_queryset().filter(id__in=obj_ids).delete()
        except ProtectedError:
            raise ValidationError('非法操作，请删除相关联数据后再试！')
        return Response(status=status.HTTP_204_NO_CONTENT)

    @atomic()
    @action(methods=['post'], detail=False, permission_classes=[IsAuthenticated], url_path='batch-update',
            url_name='batch-update')
    def batch_update(self, request):
        """批量启用、停用"""
        obj_ids = self.request.data.get('obj_ids')
        for i in obj_ids:
            try:
                instance = self.get_queryset().get(id=i)
            except Exception:
                raise ValidationError('object does not exists!')
            if instance.is_used:
                instance.is_used = False
            else:
                instance.is_used = True
            instance.save()
        return Response('ok')


def len_byte(value):
    if value is None or value == "":
        return 10
    if type(value) != int:
        if type(value) == float:
            value = str(value)
        length = len(value)
        utf8_length = len(value.encode('utf-8'))
        length = (utf8_length - length) / 2 + length
    else:
        length = len(str(value))
    return int(length)


# 数据导出通用方法 ===>服务器上生成导出文件，前端访问，通过nginx服务器返回更高效
def gen_template_response(export_fields_dict, data, file_name,
                          sheet_name='sheet1', template_file='xlsx_template/example.xlsx'):
    export_fields = list(export_fields_dict.values())
    sheet_heads = list(export_fields_dict.keys())
    excel_file = os.path.join(settings.MEDIA_ROOT, 'outfile', file_name)
    if not os.path.exists(os.path.dirname(excel_file)):
        os.makedirs(os.path.dirname(excel_file))

    # 创建一个文件对象
    wb = load_workbook(template_file)
    # 创建一个sheet对象
    sheet = wb.worksheets[0]
    if sheet_name:
        sheet.title = sheet_name
    col_width = {}
    for idx, sheet_head in enumerate(sheet_heads):
        sheet.cell(1, idx + 1).value = sheet_head
        h_w = 256 * (len_byte(sheet_head) + 2)
        col_width[sheet_head] = h_w if h_w <= 65536 else 65536

    data_row = 2
    for i in data:
        for col_num, data_key in enumerate(export_fields):
            set_value = i[data_key]
            if isinstance(set_value, str):
                set_value = cell.cell.ILLEGAL_CHARACTERS_RE.sub(r'', set_value)
            sheet.cell(data_row, col_num + 1).value = set_value
            # 设置宽度
            c_width = 256 * (len_byte(set_value) + 2) if 256 * (len_byte(set_value) + 2) <= 65536 else 65536
            if col_width[sheet_heads[col_num]] < c_width:
                col_width[sheet_heads[col_num]] = c_width
        data_row += 1
    # 设置宽度
    for col_num, data_key in enumerate(sheet_heads):
        sheet.column_dimensions[get_column_letter(col_num + 1)].width = col_width[data_key] / 256

    wb.save(excel_file)
    out_url = excel_file.replace(settings.MEDIA_ROOT, settings.MEDIA_URL, 1)
    url = parse.quote(out_url)
    return Response(data={"url": url})


def get_cur_sheet(excel_file):
    """
    获取当前工作sheet
    @param excel_file: excel模板文件
    @return: 当前工作sheet
    """
    file_name = excel_file.name
    if not file_name.split('.')[-1] in ['xls', 'xlsx', 'xlsm']:
        raise serializers.ValidationError('文件格式错误,仅支持 xls、xlsx、xlsm文件')
    try:
        data = xlrd.open_workbook(filename=None, file_contents=excel_file.read())
        cur_sheet = data.sheets()[0]
    except Exception:
        raise serializers.ValidationError('打开文件错误')
    return cur_sheet


def get_sheet_data(sheet, start_row=1):
    """
    获取excel文件所有数据
    @param start_row: 开始取数据的行数
    @param sheet:当前工作sheet
    @return: sheet列表数据
    """
    rows_num = sheet.nrows  # sheet行数
    if rows_num <= start_row:
        return []
    ret = [None] * (rows_num - start_row)
    for i in range(start_row, rows_num):
        ret[i - start_row] = sheet.row_values(i)
    return ret


def operate_record(data):
    serializer = UserOperationLogSerializer(data=data)
    if not serializer.is_valid():
        error_logger.error(f'记录操作履历出现异常: {serializer.error_messages}')
    else:
        serializer.save()


class UserFunctions(object):
    """
    针对User类进行扩展
    """

    @property
    def group_list(self):
        """
        获取用户现在所属角色列表
        :return: list
        """
        return list(self.groups.values_list('name', flat=True))

    @property
    def permissions_list(self):
        """
        获取用户所有权限id
        :return: 权限id列表
        """
        permissions = {}
        permission_ids = []
        if self.is_superuser:
            ps = Permissions.objects.filter(parent__isnull=False).order_by('parent_id', 'id').values('parent__code', 'code')
        else:
            for group in self.group_extensions.all():
                permission_ids += list(group.permissions.values_list('id', flat=True))
            permission_ids += list(self.permissions.values_list('id', flat=True))
            ps = Permissions.objects.filter(id__in=set(permission_ids), parent__isnull=False).order_by('parent_id', 'id').values('parent__code', 'code')
        for i in ps:
            if i['parent__code'] not in permissions:
                permissions[i['parent__code']] = [i['code'].split('_')[0]]
            else:
                permissions[i['parent__code']].append(i['code'].split('_')[0])
        return permissions

    def model_permission(self, model_name):
        """
        获取用户关于具体一个model的权限
        :param model_name:
        :return:
        """
        return self.get_all_permissions()

    @property
    def raw_permissions(self):
        permissions = []
        for group in self.group_extensions.all():
            group_permissions = list(group.permissions.values_list('code', flat=True))
            permissions.extend(group_permissions)
        return permissions


User.__bases__ += (UserFunctions,)
